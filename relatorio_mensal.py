import os
import mysql.connector
import smtplib
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

# === Conexão com o banco ===
conn = mysql.connector.connect(
    host='localhost',
    user='root',
    password='Mbm@20251234',
    database='sistema_os'
)
cursor = conn.cursor(dictionary=True)

# === Função para enviar email ===
def enviar_email(destinatarios, assunto, mensagem, caminhos_anexos=None):
    remetente = 'ti2@supricopygyn.com.br'
    senha = 'Mbm@2025'

    msg = MIMEMultipart()
    msg['From'] = remetente
    msg['To'] = ", ".join(destinatarios)
    msg['Subject'] = assunto
    msg.attach(MIMEText(mensagem, 'html'))

    if caminhos_anexos:
        for caminho in caminhos_anexos:
            nome_arquivo = os.path.basename(caminho)
            with open(caminho, 'rb') as file:
                parte = MIMEBase('application', 'octet-stream')
                parte.set_payload(file.read())
                encoders.encode_base64(parte)
                parte.add_header('Content-Disposition', f'attachment; filename={nome_arquivo}')
                msg.attach(parte)

    try:
        servidor = smtplib.SMTP('smtp.supricopygyn.com.br', 587)
        servidor.starttls()
        servidor.login(remetente, senha)
        servidor.sendmail(remetente, destinatarios, msg.as_string())
        servidor.quit()
        print("Relatório enviado com sucesso!")
    except Exception as e:
        print("Erro ao enviar relatório:", e)

# === Gerar relatório ===
hoje = datetime.datetime.now()
primeiro_dia = hoje.replace(day=1)
ultimo_dia = (hoje.replace(day=1) + datetime.timedelta(days=32)).replace(day=1) - datetime.timedelta(days=1)

cursor.execute("""
    SELECT os.id, u.nome AS solicitante, os.setor, os.tipo_servico,
           os.descricao, os.status, os.resolucao, os.data_criacao, os.data_fechamento
    FROM ordens_servico os
    JOIN usuarios u ON os.solicitante_id = u.id
    WHERE os.data_criacao BETWEEN %s AND %s
""", (primeiro_dia, ultimo_dia))

ordens = cursor.fetchall()

# === Criar planilha Excel com estilo ===
wb = Workbook()
ws = wb.active
ws.title = "Resumo_OS"

# Estilos
bold_font = Font(bold=True)
header_fill = PatternFill("solid", fgColor="E06666")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
border_style = Border(
    left=Side(border_style="thin"),
    right=Side(border_style="thin"),
    top=Side(border_style="thin"),
    bottom=Side(border_style="thin")
)

# Título
mes_ano = hoje.strftime('%B de %Y').capitalize()
ws.merge_cells('A1:K1')
ws['A1'] = f"Relatório de Ordens de Serviço - {mes_ano}"
ws['A1'].font = Font(size=14, bold=True)
ws['A1'].alignment = Alignment(horizontal="center")

# Subtítulo com total de OSs
ws.merge_cells('A2:K2')
ws['A2'] = f"Total de Chamados do Mês: {len(ordens)}"
ws['A2'].font = Font(size=12, bold=True)
ws['A2'].alignment = Alignment(horizontal="center")

# Cabeçalho
cabecalho = ["ID", "Solicitante", "Setor", "Tipo de Serviço", "Descrição", "Status", "Resoluçao", "Data de Criação", "Data de Fechamento", "Qtd. Anexos"]
ws.append(cabecalho)
linha_cabecalho = 3

for col_num, col_nome in enumerate(cabecalho, 1):
    celula = ws.cell(row=linha_cabecalho, column=col_num)
    celula.font = bold_font
    celula.fill = header_fill
    celula.alignment = header_alignment
    celula.border = border_style

# Dados das OS
for os_item in ordens:
    os_id = os_item['id']
    cursor.execute("SELECT nome_arquivo FROM anexos_os WHERE os_id = %s", (os_id,))
    anexos = cursor.fetchall()
    qtd_anexos = len(anexos)

    linha = [
        os_id,
        os_item['solicitante'],
        os_item['setor'],
        os_item['tipo_servico'],
        os_item['descricao'],
        os_item['status'],
        os_item['resolucao'],
        os_item['data_criacao'].strftime('%d/%m/%Y %H:%M'),
        os_item['data_fechamento'].strftime('%d/%m/%Y %H:%M') if os_item['data_fechamento'] else '-',
        qtd_anexos
    ]
    ws.append(linha)

# Aplicar borda e alinhamento nos dados
for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=11):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border_style

# Ajuste de largura de coluna
for col in ws.columns:
    max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
    col_letter = get_column_letter(col[0].column)
    ws.column_dimensions[col_letter].width = max(max_length + 2, 15)

# Abas separadas para OSs com anexos
for os_item in ordens:
    os_id = os_item['id']
    cursor.execute("SELECT nome_arquivo FROM anexos_os WHERE os_id = %s", (os_id,))
    anexos = cursor.fetchall()

    if anexos:
        aba = wb.create_sheet(f"OS_{os_id}")
        aba.append(["Nome do Arquivo", "Caminho", "Link"])
        for cell in aba[1]:
            cell.font = bold_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border_style

        for anexo in anexos:
            nome = anexo['nome_arquivo']
            caminho = os.path.join('uploads', nome)
            link = f'=HYPERLINK("{caminho}", "Abrir")'
            aba.append([nome, caminho, link])

        # Estilizar linhas
        for row in aba.iter_rows(min_row=2, max_row=aba.max_row, min_col=1, max_col=3):
            for cell in row:
                cell.border = border_style
                cell.alignment = Alignment(vertical="top")

        # Ajustar largura
        for col in aba.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            col_letter = get_column_letter(col[0].column)
            aba.column_dimensions[col_letter].width = max(max_length + 2, 20)

# === Salva o Excel ===
os.makedirs("relatorios", exist_ok=True)
caminho_excel = os.path.join("relatorios", f"relatorio_os_{hoje.strftime('%Y_%m')}.xlsx")
wb.save(caminho_excel)

# === Envia o e-mail ===
enviar_email(
    ["depto.ti1@mbmcopy.com.br", "depto.ti2@mbmcopy.com.br", "paulo.faraone@mbmcopy.com.br"],
    f"Relatório Mensal de OS - {hoje.strftime('%B %Y')}",
    f"""
    <p>Prezados,</p>
    <p>Segue em anexo o <b>relatório mensal das Ordens de Serviço</b> com o total de <b>{len(ordens)} chamados</b>  referente ao período de 
    <b>{primeiro_dia.strftime('%d/%m/%Y')}</b> a <b>{ultimo_dia.strftime('%d/%m/%Y')}</b>.</p>
    <p>As OSs estão organizadas em uma aba de resumo e anexos detalhados em abas separadas por OS.</p>
    <p>Atenciosamente,<br>TI MBMCOPY</p>
    """,
    [caminho_excel]
)

cursor.close()
conn.close()
